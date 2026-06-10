import os
import re
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
XLSX_PATH = BASE_DIR / "salidas" / "Ranking de procesamiento por ODPE.xlsx"
DESCARGAS_DIR = BASE_DIR / "insumos" / "descargas_modulo"

TITLE = "Elección presidencial de segunda vuelta Perú 2026"
SUBTITLE = "Avance de la contabilización de actas por ODPE a nivel nacional"
FALLBACK_UPDATE = "Actualización: 8 de junio – 1:30pm"
SOURCE = "Fuente: Sistema de presentación de resultados – Módulo Especializado - ONPE."


def infer_update_label():
    if os.environ.get("ONPE_SV_UPDATE_LABEL"):
        return os.environ["ONPE_SV_UPDATE_LABEL"]

    month_names = {
        "01": "enero",
        "02": "febrero",
        "03": "marzo",
        "04": "abril",
        "05": "mayo",
        "06": "junio",
        "07": "julio",
        "08": "agosto",
        "09": "septiembre",
        "10": "octubre",
        "11": "noviembre",
        "12": "diciembre",
    }
    csvs = sorted(DESCARGAS_DIR.glob("*.csv"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not csvs:
        return FALLBACK_UPDATE

    match = re.search(r"_(\d{4})-(\d{2})-(\d{2})_(\d{2})-(\d{2})-(AM|PM)_", csvs[0].name, re.I)
    if not match:
        return FALLBACK_UPDATE

    _, month, day, hour, minute, ampm = match.groups()
    hour_num = int(hour) or 12
    return f"Actualización: {int(day)} de {month_names.get(month, month)} – {hour_num}:{minute}{ampm.lower()}"


def main():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"No existe el archivo: {XLSX_PATH}")

    wb = load_workbook(XLSX_PATH)
    ws = wb["Ranking ODPE"]

    ws["A1"] = TITLE
    ws["A2"] = SUBTITLE
    ws["A3"] = infer_update_label()
    ws.cell(ws.max_row, 1).value = SOURCE
    ws.freeze_panes = ws["A6"]

    wb.save(XLSX_PATH)
    print(f"Excel finalizado: {XLSX_PATH}")


if __name__ == "__main__":
    main()
